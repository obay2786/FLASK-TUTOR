import os
import re
from flask import Blueprint, render_template, request, flash, jsonify, send_from_directory, redirect,current_app,url_for
from flask_login import login_required, current_user
import datetime
from .models import Badge, Visitor,User, Permit, Location, Transaksi
from . import db
import json
import base64
from sqlalchemy import asc, desc,text,func
import requests
from PIL import Image, ImageOps
from io import BytesIO
from .mailr import kirimEmail

from openpyxl import load_workbook
views = Blueprint('views', __name__)
ROWS_PER_PAGE = 10


def saveB(photo):
  
    img = Image.open(photo)
    img = ImageOps.exif_transpose(img)
    output_size = (600, 600)
    img.thumbnail(output_size)
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    img_str = base64.b64encode(buffered.getvalue())
    
    return img_str

def saveGambar(photo):
  
    img = Image.open(photo)
    img = ImageOps.exif_transpose(img)
    output_size = (1300, 1300)
    img.thumbnail(output_size)
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    img_str = base64.b64encode(buffered.getvalue())
    
    return img_str
def qrGen(id,nik,nama):
    qr = f'https://chart.apis.google.com/chart?chl={id}:{nik}&chs=150x150&cht=qr&chld=H%7C0'
    
    return {nama:qr}
    

@views.route('/gambarpermit', methods=['GET', 'POST'])
@login_required
def gambarpermit():
    if request.method == 'POST':
        print(request.form)
        gambarPermit = saveGambar(request.files['file'])
        permitId = request.form.get('idPermit')
        permit = Permit.query.filter_by(id=permitId).first()
        permit.UploadPermit = gambarPermit
        permit.status = 'approved'
        db.session.commit()
        return jsonify({})


@views.route('/', methods=['GET', 'POST'])
@login_required
def home():
    if request.method == 'POST':
        if request.form.get('formEdit') == 'uploadGambarApproval':
            print(request.form)
            gambarPermit = saveGambar(request.files['photo'])
            permitId = request.form.get('idPermit')
            permit = Permit.query.filter_by(id=permitId).first()
            permit.UploadPermit = gambarPermit
            permit.status = 'approved'
            db.session.commit()

        return redirect(url_for('views.home'))
    else:
        page = request.args.get('page', 1, type=int)
        host = current_user.firstName + ':' +current_user.empID
        permit = Permit.query.filter_by(host=host, status='waitinghost').order_by(text('id desc')).paginate(page=page, per_page=ROWS_PER_PAGE)
        transaksi = Transaksi.query.order_by(text('id desc')).paginate(page=page, per_page=ROWS_PER_PAGE)
        location = Location.query.order_by(Location.id).all()
        # status = Permit.query.order_by(text('status')).paginate(page=page, per_page=ROWS_PER_PAGE)
        
        return render_template("home.html", user=current_user, permit=permit,location=location,transaksi=transaksi)

@views.route('/waiting', methods=['GET', 'POST'])
@login_required
def waiting():
    startDate = datetime.datetime.now()
    endDate = datetime.datetime.now()
    transaksi = Transaksi.query.filter(text(f'timeCheckin BETWEEN \'{startDate.strftime("%Y/%m/%d")} 00:00:00\' AND \'{endDate.strftime("%Y/%m/%d")} 23:59:59\'')).order_by(text('id desc')).all()

    return render_template("waiting.html", user=current_user, transaksi=transaksi)

@views.route('/waitinglist', methods=['GET', 'POST'])
#@login_required
def waitinglist():
    if request.method == 'POST':
        ts = json.loads(request.data)
        tsId = ts['tsid']
        transaksi = Transaksi.query.filter_by(id=tsId).first()
        transaksi.status = 'checkin'
        db.session.commit()
        return jsonify({})


    else:
        transaksi = Transaksi.query.filter_by(status= 'waiting').order_by(text('id desc')).all()
        
        data = []
        dataDict ={}
        for i in transaksi:
            dataDict['id'] = i.id
            dataDict['nama'] = i.namaVisitor
            dataDict['nik'] = i.nik
            dataDict['company'] = i.vendor
            dataDict['location'] = i.location
            dataDict['photo'] = i.photo
            dataDict['badge'] = i.badge
            
            data.append(dataDict.copy())
        #print(data)
        return jsonify(data)


@views.route('/history', methods=['GET', 'POST'])
@login_required
def history():
    page = request.args.get('page', 1, type=int)
    host = current_user.firstName + ':' +current_user.empID
    permit = Permit.query.filter_by(host=host,status='approved').order_by(text('id desc')).paginate(page=page, per_page=ROWS_PER_PAGE)
    return render_template("history.html", user=current_user, permit=permit)



@views.route('/approval', methods=['GET', 'POST'])
@login_required
def approval():
    page = request.args.get('page', 1, type=int)
    permitWorking = Permit.query.filter_by(purpose='WORKING',status='waitingadmin').order_by(text('id desc')).paginate(page=page)
    permitOvertime = Permit.query.filter_by(purpose='OVERTIME',status='waitingadmin').order_by(text('id desc')).paginate(page=page)
    return render_template("approval.html", user=current_user, permitWorking=permitWorking, permitOvertime=permitOvertime)


@views.route('/visitor',methods=['GET', 'POST'])
@login_required
def visitor():
    if request.method == 'POST':
        if request.form.get('formEdit') == 'edit':
            print(request.form)
            data ={}
            
            data['id'] = request.form.get('id')
            data['nama'] = request.form.get('nama')
            data['nik'] = request.form.get('nik')
            data['company'] = request.form.get('company')
            
            print(data)
            dataQ = Visitor.query.filter_by(id=data['id']).first()
            dataQ.nama = data['nama']
            dataQ.nik = data['nik']
            dataQ.namaVendor = data['company']
           
            db.session.commit()
            print('okeeeee')
       
        if request.form.get('formEdit') == 'editPhoto':
            print(request.form)
            data ={}
            data['id'] = request.form.get('id')
            data['photo'] = saveB(request.files['photo'])
            
            
            print(data)
            dataQ = Visitor.query.filter_by(id=data['id']).first()
            dataQ.photo = data['photo']
            
            db.session.commit()
            print('okeeeee')
    page = request.args.get('page', 1, type=int)
    print(page)
    #data = Visitor.query.paginate(page=page, per_page=ROWS_PER_PAGE)
    data = Visitor.query.order_by(text('id desc')).paginate(page=page, per_page=ROWS_PER_PAGE)
    #data.reverse()
    #data.items.reverse()
    return render_template("adminvisitordata.html", user=current_user,data=data)



@views.route('/report',methods=['GET', 'POST'])
@login_required
def report():
    if current_user.role == 'Admin' or current_user.role == 'Security':
        # data = json.loads(request.data)
        # by = data['by']
        # transaksi = Transaksi.query.order_by(text("id desc")).all()
        ID = None
        host = None
        purpose = None
        company = None
        my_filters = {'id':ID, 'host':host, 'purpose':purpose, 'vendor':company}
        filterz = {}
        for i,j in my_filters.items():
            if j != None:
                filterz.update({i:j})
        
        transaksi = Transaksi.query.order_by(text('id desc')).filter_by(**filterz).all()                  #filter(text(f'timeCheckin BETWEEN \'{startDate.strftime("%Y/%m/%d")} 00:00:00\' AND \'{endDate.strftime("%Y/%m/%d")} 23:59:59\''))


        return render_template("report.html", user=current_user, transaksi=transaksi)
    else:
        return render_template("login.html", user=current_user)

@views.route('/nama')
@login_required
def namadelete_note():
    nama = {"nama":"novel martin harianto"}
    return jsonify(nama)

@views.route("/assets/<path:path>")
def css(path):
    return send_from_directory('assets', path)

@views.route("/kiosk/<path:path>")
def kiosk(path):
    return send_from_directory('kiosk', path)

@views.route("/kiosk/")
def kioskindex():
    return redirect("./index.html", code=302)

@views.route("/checked/<path:path>")
def cek(path):
    return send_from_directory('checked', path)

@views.route("/check-in")
def kcheckin():
    return redirect("./checked/check-in.html", code=302)

@views.route("/check-out")
def kcheckout():
    return redirect("./checked/check-out.html", code=302)

@views.route('/delstaff', methods=['POST'])
def delStaff():
    data = json.loads(request.data)
    staffID= data['id']
    print(staffID)
    staff = User.query.filter_by(id=staffID).first()
    if staff:
        db.session.delete(staff)
        db.session.commit()
    print(staff.id)
    return jsonify({})

@views.route('/getstaff', methods=['POST'])
def getStaff():
    data = json.loads(request.data)
    staffID = data['id']
    print(staffID)
    staff = User.query.filter_by(id=staffID).first()
    data = {}
    data['id'] = staff.id
    data['userName'] = staff.userName
    data['firstName'] = staff.firstName
    data['email'] = staff.email
    data['empID'] = staff.empID
    data['badgeID'] = staff.badgeID
    data['role'] = staff.role
    data['depart'] = staff.depart
    data['photo'] = staff.photo
    return jsonify(data)

@views.route('/getstaffco', methods=['POST'])
def getStaffco():
    data = json.loads(request.data)
    staffID = data['badgeID']
    print(staffID)
    staff = User.query.filter_by(badgeID=staffID).first()
    # print('ooookkkkkkiiii'+staff)

    if staff:
        data = {}
        data['id'] = staff.id
        data['userName'] = staff.userName
        data['firstName'] = staff.firstName
        data['email'] = staff.email
        data['empID'] = staff.empID
        data['badgeID'] = staff.badgeID
        data['role'] = staff.role
        data['depart'] = staff.depart
        data['photo'] = staff.photo
        return jsonify(data)
    else:
        return jsonify({"empID":"tidak"})

@views.route('/getvisitor', methods=['POST'])
def getVisitor():
    data = json.loads(request.data)
    nik = data['nik']
    print(nik)
    visitor = Visitor.query.filter_by(nik=nik).first()
    print(visitor)
    data = {}
    data['id'] = visitor.id
    data['nik'] = visitor.nik
    data['nama'] = visitor.nama

    data['company'] = visitor.namaVendor
    data['photo'] = visitor.photo
    
    return jsonify(data)



@views.route('/delvisitor', methods=['POST'])
def delVisitor():
    data = json.loads(request.data)
    visitorID= data['id']
    print(visitorID)
    visitorD = Visitor.query.filter_by(id=visitorID).first()
    if visitorD:
        db.session.delete(visitorD)
        db.session.commit()
    
    return jsonify({})

@views.route('/permitdetail', methods=['POST'])
def getPermitdetail():
    data = json.loads(request.data)
    permitId = data['id']
    # print(id)data[id]
    permit = Permit.query.filter_by(id=permitId).first()
    # print(visitor)
    data = {}
    data['id'] = permit.id
    data['date'] = permit.subDate
    data['vendor'] = permit.namaVendor
    data['startDate'] = permit.startDate
    data['endDate'] = permit.endDate
    data['purpose'] = permit.purpose
    data['anggota'] = json.loads(permit.anggota)
    data['permitNo'] = permit.permitNo
    data['desk'] = permit.desk
    data['location'] = permit.location
    data['supplyBarang'] = permit.supplyBarang
    data['email'] = permit.email
    data['host'] = permit.host
    data['bawaBarang'] = permit.bawaBarang
    if permit.bawaBarang == 'TIDAK':
        data['barangBawaan'] = ''
    else:
        data['barangBawaan'] = json.loads(permit.barangBawaan)
    data['sign'] = permit.sign
    data['buttongenerate'] = ""
    if "'tidak'" in str(data['anggota']):
        data['buttongenerate'] = 'disable'
    else:
        data['buttongenerate'] = 'enable' 

    data['uploadGambar'] = permit.UploadPermit

    print('ini hasil dari' + data['buttongenerate'])
    return jsonify(data)

@views.route('/updatepermitlocation', methods=['POST'])
@login_required
def updatepermitlocation():
    data = json.loads(request.data)
    print(data)
    permitId = data['id']
    permitLoc = data['location']
    # print(id)
    permit = Permit.query.filter_by(id=permitId).first()
    
    # print(visitor)
   
    permit.location = permitLoc
    db.session.commit()
    

    data = {}
    return jsonify(data)


@views.route('/genxls', methods=['POST'])

def genxls():
    data = json.loads(request.data)
    print(data)
    permitId = data['id']
    permit = Permit.query.filter_by(id=permitId).first()
    

    sourceFile = r'VisitorApprovalBT.xlsx';
    urlFolder = os.path.join(current_app.root_path,'static',sourceFile)

    wb = load_workbook(urlFolder);

    sheet = wb['Visitor Approval'];

    #hapus
    sheet['C5'] = "" 
    sheet['C6'] = ""
    sheet['C7'] = ""
    sheet['C9'] = ""
    sheet['C10'] = ""
    sheet['E10'] = ""
    sheet['A11'] = ""
    sheet['A11'] = ""
    for a in range(7,17):
        sheet[f'H{a}'] = ""
    for b in range(21,31):
        sheet[f'D{b}'] = ""
        sheet[f'E{b}'] = ""
    for c in range(36,66):
        sheet[f'D{c}'] = ""


    #isi
    anggota = json.loads(permit.anggota)
    sheet['C5'] = anggota[0]['Nama']  #nama Anggota pertama 
    sheet['C6'] = permit.location       #location
    sheet['C7'] = permit.namaVendor #namavendor
    sheet['C9'] = str(permit.startDate)[0:10] + ' - ' +  str(permit.endDate)[0:10]#startdata dan enddate
    sheet['C10'] = permit.startDate 
    sheet['E10'] = permit.endDate
    sheet['A11'] = permit.desk
    # sheet['A11'] = permit.desk
    
    
    for i,j in enumerate(anggota, start=7):
        sheet[f'H{i}'] = j['Nama']
   
    if permit.bawaBarang == 'ya':
        barang = json.loads(permit.barangBawaan)
        sumber = {21:'Personal Computer / Laptop',22:'Camera (Digital or analogue)',23:'Mobilephone with camera / video',24:'Tablet with camera / video',25:'Digital Video Recorder',26:'Thumbdrive / Pendrive storage unit',27:'Memory Cards (SD/CF/MMC etc.)',28:'Audio Tape Recorder',29:'CDRW / CDR / HDD',30:'Others (pls state)'}
        hasil = set()
        for i in barang:
            for key, value in sumber.items():
                if i['Jenis Media'] == value:
                    # hasil.add(key)
                    sheet[f'D{key}'] = str('✅')
                    sheet[f'E{key}'] = i['Tujuan']
    else:
        for b in range(21,31):
            sheet[f'D{b}'] = ""
            sheet[f'E{b}'] = ""

    ws = wb['Visitor Approval']
    locations = []
    for row in ws.iter_rows(min_row=36, min_col=3, max_col=3, max_row=65, values_only=True):
        x = locations.append(row[0])

    detaillocations = {}
    for i, v in enumerate(locations, start=36):
        detaillocations[i] = v
    for key, values in detaillocations.items():
        if permit.location == values:
            sheet[f'D{key}'] = str('✅')

    # buffer = BytesIO()
    wb.save(urlFolder)

    permit.statusGenerate = 'ok'
    db.session.commit()
    
    

    return jsonify({})



@views.route('/kirimemaildecline', methods=['POST'])
def kirimEmailDecline():
    data = json.loads(request.data)
    print(data)
    permitId = data['id']
    permit = Permit.query.filter_by(id=permitId).first()
    email = permit.email
    subject = 'Permit Reject'
    body = data['body']
    kirimEmail(email,subject,body)
    db.session.delete(permit)
    db.session.commit()
    return jsonify({})

@views.route('/kirimemailapprove', methods=['POST'])
def kirimEmailApprove():
    data = json.loads(request.data)
    print(data)
    
    id = data['id']
    permit = Permit.query.filter_by(id=id).first()
    anggota = json.loads(permit.anggota)
    listQR = []
    for person in anggota:
        listQR.append(qrGen(id,person['NIK'],person['Nama']))

    print(listQR)
    
    email = permit.email
    subject = 'Permit Accepted'
    tableQR = []
    
    for QR in listQR:
        for key in QR:
            tableQR.append(f"""<td class="column" style="mso-table-lspace:0;mso-table-rspace:0;font-weight:400;text-align:left;vertical-align:top;border-bottom:2px solid #fff;border-left:2px solid #fff;border-right:2px solid #fff;border-top:2px solid #fff" width="33.333333333333336%">
                                                    <table class="image_block" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0" width="100%" cellspacing="0" cellpadding="0" border="0">
                                                        <tbody><tr>
                                                            <td style="width:100%;padding-right:0;padding-left:0;padding-top:5px">
                                                                <div style="line-height:10px" align="center"><img src="{QR[key]}" style="display:block;height:auto;border:0;width:183px;max-width:100%" alt="Hospital Icon" title="Hospital Icon" width="183"></div>
                                                            </td>
                                                        </tr>
                                                    </tbody></table>
                                                    <table class="text_block" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0;word-break:break-word" width="100%" cellspacing="0" cellpadding="10" border="0">
                                                        <tbody><tr>
                                                            <td>
                                                                <div style="font-family:Tahoma,Verdana,sans-serif">
                                                                    <div style="font-size:12px;font-family:Lato,Tahoma,Verdana,Segoe,sans-serif;mso-line-height-alt:14.399999999999999px;color:#087200;line-height:1.2">
                                                                        <p style="margin:0;font-size:16px;text-align:center">
                                                                            <strong>{key}</strong></p>
                                                                    </div>
                                                                </div>
                                                            </td>
                                                        </tr>
                                                    </tbody></table>
                                                    <table class="divider_block" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0" width="100%" cellspacing="0" cellpadding="10" border="0">
                                                        <tbody><tr>
                                                            <td>
                                                                <div align="center">
                                                                    <table role="presentation" style="mso-table-lspace:0;mso-table-rspace:0" width="100%" cellspacing="0" cellpadding="0" border="0">
                                                                        <tbody><tr>
                                                                            <td class="divider_inner" style="font-size:1px;line-height:1px;border-top:1px solid #bbb">
                                                                                <span> </span></td>
                                                                        </tr>
                                                                    </tbody></table>
                                                                </div>
                                                            </td>
                                                        </tr>
                                                    </tbody></table>
                                                    <table class="text_block" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0;word-break:break-word" width="100%" cellspacing="0" cellpadding="0" border="0">
                                                        <tbody><tr>
                                                            <td style="padding-bottom:15px;padding-left:10px;padding-right:10px;padding-top:10px">
                                                                <div style="font-family:Tahoma,Verdana,sans-serif">
                                                                    <div style="font-size:12px;font-family:Lato,Tahoma,Verdana,Segoe,sans-serif;mso-line-height-alt:18px;color:#000;line-height:1.5">
                                                                        <p style="margin:0;font-size:14px;text-align:center;mso-line-height-alt:24px">
                                                                            <span style="font-size:16px">Lorem ipsum
                                                                                dolor sit amet,</span><br><span style="font-size:16px">consectetur
                                                                                adipiscing elit</span></p>
                                                                    </div>
                                                                </div>
                                                            </td>
                                                        </tr>
                                                    </tbody></table>
                                                </td>""")
    #print(tableQR[0])
    rowPertama = ''
    rowKedua = ''
    rowKetiga = ''
    rowKeempat = ''
    rowKelima = ''
    rowKeenam = ''
    rowKetujuh = ''

    listQrMail = []
    for t in range(len(tableQR)):
        if t < 3:
            rowPertama += tableQR[t]
            
        elif t < 6:
            rowKedua += tableQR[t]
        elif t < 9:
            rowKetiga += tableQR[t]
        elif t < 12:
            rowKeempat += tableQR[t]
        elif t < 15:
            rowKelima += tableQR[t]
        elif t < 18:
            rowKeenam += tableQR[t]
        else:
            rowKetujuh += tableQR[t]

    #print(rowPertama)
    tableRow =[rowPertama,rowKedua,rowKetiga,rowKeempat,rowKelima,rowKeenam,rowKetujuh]
    tableRowFinish = []
    for i in tableRow:
        tableRowFinish.append(f"""<table class="row row-4" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0" width="100%" cellspacing="0" cellpadding="0" border="0" align="center">
                                <tbody>
                                    <tr>
                                        <td>
                                            <table class="row-content stack" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0;background-color:#fff;color:#000;width:680px" width="680" cellspacing="0" cellpadding="0" border="0" align="center">
                                                <tbody>
                                                    <tr>
                                                    {i}
                                                    </tr>
                                                </tbody>
                                                
                                            </table>
                                        </td>
                                    </tr>
                                </tbody>
                            </table>""")
        
    bodyFinish = f"""<!DOCTYPE html>
<html xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office" lang="en"><head>
    <title></title>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width,initial-scale=1">
    <!--[if mso]><xml><o:OfficeDocumentSettings><o:PixelsPerInch>96</o:PixelsPerInch><o:AllowPNG/></o:OfficeDocumentSettings></xml><![endif]-->
    <!--[if !mso]><!-->
    <link href="https://fonts.googleapis.com/css?family=Lato" rel="stylesheet" type="text/css">
    <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@700&amp;display=swap" rel="stylesheet" type="text/css">
    <link href="https://fonts.googleapis.com/css?family=Permanent+Marker" rel="stylesheet" type="text/css">
    <!--<![endif]-->
    <style>
        * {{
            box-sizing: border-box
        }}

        body {{
            margin: 0;
            padding: 0
        }}

        a[x-apple-data-detectors] {{
            color: inherit !important;
            text-decoration: inherit !important
        }}

        #MessageViewBody a {{
            color: inherit;
            text-decoration: none
        }}

        p {{
            line-height: inherit
        }}

        @media (max-width:700px) {{
            .icons-inner {{
                text-align: center
            }}

            .icons-inner td {{
                margin: 0 auto
            }}

            .row-content {{
                width: 100% !important
            }}

            .stack .column {{
                width: 100%;
                display: block
            }}
        }}
    </style>
</head>

<body style="background-color:#fff7f7;margin:0;padding:0;-webkit-text-size-adjust:none;text-size-adjust:none">
    <table class="nl-container" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0;background-color:#fff7f7" width="100%" cellspacing="0" cellpadding="0" border="0">
        <tbody>
            <tr>
                <td>
                    <table class="row row-1" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0" width="100%" cellspacing="0" cellpadding="0" border="0" align="center">
                        <tbody>
                            <tr>
                                <td>
                                    <table class="row-content stack" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0;color:#000;width:680px" width="680" cellspacing="0" cellpadding="0" border="0" align="center">
                                        <tbody>
                                            <tr>
                                                <td class="column" style="mso-table-lspace:0;mso-table-rspace:0;font-weight:400;text-align:left;vertical-align:top;padding-top:5px;padding-bottom:5px;border-top:0;border-right:0;border-bottom:0;border-left:0" width="100%">
                                                    <div class="spacer_block" style="height:40px;line-height:40px;font-size:1px"> </div>
                                                </td>
                                            </tr>
                                        </tbody>
                                    </table>
                                </td>
                            </tr>
                        </tbody>
                    </table>
                    <table class="row row-2" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0" width="100%" cellspacing="0" cellpadding="0" border="0" align="center">
                        <tbody>
                            <tr>
                                <td>
                                    <table class="row-content stack" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0;background-color:#fff;color:#000;width:680px" width="680" cellspacing="0" cellpadding="0" border="0" align="center">
                                        <tbody>
                                            <tr>
                                                <td class="column" style="mso-table-lspace:0;mso-table-rspace:0;font-weight:400;text-align:left;vertical-align:top;padding-top:0;padding-bottom:0;border-top:0;border-right:0;border-bottom:0;border-left:0" width="100%">
                                                    <table class="image_block" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0" width="100%" cellspacing="0" cellpadding="0" border="0">
                                                        <tbody><tr>
                                                            <td style="width:100%;padding-right:0;padding-left:0;padding-top:40px">
                                                                <div style="line-height:10px" align="center"><img src="https://d1oco4z2z1fhwp.cloudfront.net/templates/default/5126/logo.png" style="display:block;height:auto;border:0;width:28px;max-width:100%" alt="Red Ribbon Logo Placeholder" title="Red Ribbon Logo Placeholder" width="28"></div>
                                                            </td>
                                                        </tr>
                                                    </tbody></table>
                                                    <table class="text_block" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0;word-break:break-word" width="100%" cellspacing="0" cellpadding="0" border="0">
                                                        <tbody><tr>
                                                            <td style="padding-bottom:20px;padding-left:15px;padding-right:15px;padding-top:20px">
                                                                <div style="font-family:sans-serif">
                                                                    <div style="font-size:14px;font-family:Poppins,sans-serif;mso-line-height-alt:16.8px;color:#000;line-height:1.2">
                                                                        <p style="margin:0;font-size:30px;text-align:center">
                                                                            <span style="font-size:50px"><span style=""><strong>PERMIT
                                                                                    </strong></span><span style=""><strong><span style="color:#fe000b">PANASONIC
                                                                                        </span>BATAM</strong></span></span>
                                                                        </p>
                                                                    </div>
                                                                </div>
                                                            </td>
                                                        </tr>
                                                    </tbody></table>
                                                    <table class="text_block" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0;word-break:break-word" width="100%" cellspacing="0" cellpadding="0" border="0">
                                                        <tbody><tr>
                                                            <td style="padding-bottom:10px;padding-left:15px;padding-right:15px;padding-top:10px">
                                                                <div style="font-family:Tahoma,Verdana,sans-serif">
                                                                    <div style="font-size:14px;font-family:Lato,Tahoma,Verdana,Segoe,sans-serif;mso-line-height-alt:21px;color:#000;line-height:1.5">
                                                                        <p style="margin:0;text-align:center;font-size:16px;mso-line-height-alt:24px">
                                                                            <span style="font-size:16px">Lorem ipsum
                                                                                dolor sit amet, consectetur adipiscing
                                                                                elit. Nunc at ornare dolor.</span></p>
                                                                        <p style="margin:0;text-align:center;font-size:16px;mso-line-height-alt:24px">
                                                                            <span style="font-size:16px">Aliquet lectus
                                                                                egetullamcorper sollicitudin.&nbsp;
                                                                                Fusce in ultricies velit, accommodo
                                                                                quam.</span></p>
                                                                        <p style="margin:0;text-align:center;font-size:16px;mso-line-height-alt:24px">
                                                                            <span style="font-size:16px">Class aptent
                                                                                taciti sociosqu ad.</span></p>
                                                                        <p style="margin:0;text-align:center;font-size:16px;mso-line-height-alt:21px">
                                                                            &nbsp;</p>
                                                                    </div>
                                                                </div>
                                                            </td>
                                                        </tr>
                                                    </tbody></table>
                                                </td>
                                            </tr>
                                        </tbody>
                                    </table>
                                </td>
                            </tr>
                        </tbody>
                    </table>
                    
                    {"".join(tableRowFinish)}
                    
                    
                    
                    <table class="row row-8" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0" width="100%" cellspacing="0" cellpadding="0" border="0" align="center">
                        <tbody>
                            <tr>
                                <td>
                                    <table class="row-content stack" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0;background-color:#fe000b;color:#000;width:680px" width="680" cellspacing="0" cellpadding="0" border="0" align="center">
                                        <tbody>
                                            <tr>
                                                <td class="column" style="mso-table-lspace:0;mso-table-rspace:0;font-weight:400;text-align:left;vertical-align:top;padding-top:5px;padding-bottom:5px;border-top:0;border-right:0;border-bottom:0;border-left:0" width="100%">
                                                    <table class="social_block" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0" width="100%" cellspacing="0" cellpadding="0" border="0">
                                                        <tbody><tr>
                                                            <td style="padding-bottom:10px;padding-left:10px;padding-right:10px;padding-top:30px;text-align:center">
                                                                <table class="social-table" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0" width="144px" cellspacing="0" cellpadding="0" border="0" align="center">
                                                                    <tbody><tr>
                                                                        <td style="padding:0 2px 0 2px"><a href="https://www.facebook.com/" target="_blank"><img src="https://app-rsrc.getbee.io/public/resources/social-networks-icon-sets/t-only-logo-white/facebook@2x.png" alt="Facebook" title="facebook" style="display:block;height:auto;border:0" width="32" height="32"></a>
                                                                        </td>
                                                                        <td style="padding:0 2px 0 2px"><a href="https://www.twitter.com/" target="_blank"><img src="https://app-rsrc.getbee.io/public/resources/social-networks-icon-sets/t-only-logo-white/twitter@2x.png" alt="Twitter" title="twitter" style="display:block;height:auto;border:0" width="32" height="32"></a>
                                                                        </td>
                                                                        <td style="padding:0 2px 0 2px"><a href="https://www.linkedin.com/" target="_blank"><img src="https://app-rsrc.getbee.io/public/resources/social-networks-icon-sets/t-only-logo-white/linkedin@2x.png" alt="Linkedin" title="linkedin" style="display:block;height:auto;border:0" width="32" height="32"></a>
                                                                        </td>
                                                                        <td style="padding:0 2px 0 2px"><a href="https://www.instagram.com/" target="_blank"><img src="https://app-rsrc.getbee.io/public/resources/social-networks-icon-sets/t-only-logo-white/instagram@2x.png" alt="Instagram" title="instagram" style="display:block;height:auto;border:0" width="32" height="32"></a>
                                                                        </td>
                                                                    </tr>
                                                                </tbody></table>
                                                            </td>
                                                        </tr>
                                                    </tbody></table>
                                                    <table class="menu_block" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0" width="100%" cellspacing="0" cellpadding="0" border="0">
                                                        <tbody><tr>
                                                            <td style="color:#fff;font-family:Lato,Tahoma,Verdana,Segoe,sans-serif;font-size:12px;letter-spacing:1px;text-align:center;padding-top:20px">
                                                                <table role="presentation" style="mso-table-lspace:0;mso-table-rspace:0" width="100%" cellspacing="0" cellpadding="0" border="0">
                                                                    <tbody><tr>
                                                                        <td style="text-align:center;font-size:0">
                                                                            <div class="menu-links">
                                                                                <!--[if mso]>
    <table role="presentation" border="0" cellpadding="0" cellspacing="0" align="center" style="">
    <tr>
    <td style="padding-top:5px;padding-right:5px;padding-bottom:5px;padding-left:5px">
    <![endif]--><a href="wwwexample.com" style="padding-top:5px;padding-bottom:5px;padding-left:5px;padding-right:5px;display:inline-block;color:#fff;font-family:Lato,Tahoma,Verdana,Segoe,sans-serif;font-size:12px;text-decoration:none;letter-spacing:1px">ARCHIVE</a>
                                                                                <!--[if mso]></td><td><![endif]--><span class="sep" style="font-size:12px;font-family:Lato,Tahoma,Verdana,Segoe,sans-serif;color:#fff">|</span>
                                                                                <!--[if mso]></td><![endif]-->
                                                                                <!--[if mso]></td><td style="padding-top:5px;padding-right:5px;padding-bottom:5px;padding-left:5px"><![endif]--><a href="www.example.com" style="padding-top:5px;padding-bottom:5px;padding-left:5px;padding-right:5px;display:inline-block;color:#fff;font-family:Lato,Tahoma,Verdana,Segoe,sans-serif;font-size:12px;text-decoration:none;letter-spacing:1px">SUBSCRIBE</a>
                                                                                <!--[if mso]></td></tr></table><![endif]-->
                                                                            </div>
                                                                        </td>
                                                                    </tr>
                                                                </tbody></table>
                                                            </td>
                                                        </tr>
                                                    </tbody></table>
                                                    <table class="text_block" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0;word-break:break-word" width="100%" cellspacing="0" cellpadding="0" border="0">
                                                        <tbody><tr>
                                                            <td style="padding-bottom:30px;padding-left:15px;padding-right:25px;padding-top:30px">
                                                                <div style="font-family:Tahoma,Verdana,sans-serif">
                                                                    <div style="font-size:12px;font-family:Lato,Tahoma,Verdana,Segoe,sans-serif;mso-line-height-alt:18px;color:#fff;line-height:1.5">
                                                                        <p style="margin:0;font-size:14px;text-align:center;mso-line-height-alt:18px;letter-spacing:normal">
                                                                            <span style="font-size:12px">This email was
                                                                                intended for you. &nbsp;<a href="http://www.example.com" target="_blank" style="text-decoration:underline;color:#3d3d3d" rel="noopener">Unsubscribe</a> from
                                                                                daily digest messages,</span></p>
                                                                        <p style="margin:0;font-size:14px;text-align:center;mso-line-height-alt:18px;letter-spacing:normal">
                                                                            <span style="font-size:12px">or visit your
                                                                            </span><span style="font-size:12px">settings
                                                                                to manage what emails to send
                                                                                you.</span></p>
                                                                    </div>
                                                                </div>
                                                            </td>
                                                        </tr>
                                                    </tbody></table>
                                                </td>
                                            </tr>
                                        </tbody>
                                    </table>
                                </td>
                            </tr>
                        </tbody>
                    </table>
                    <table class="row row-9" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0" width="100%" cellspacing="0" cellpadding="0" border="0" align="center">
                        <tbody>
                            <tr>
                                <td>
                                    <table class="row-content stack" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0;color:#000;width:680px" width="680" cellspacing="0" cellpadding="0" border="0" align="center">
                                        <tbody>
                                            <tr>
                                                <td class="column" style="mso-table-lspace:0;mso-table-rspace:0;font-weight:400;text-align:left;vertical-align:top;padding-top:5px;padding-bottom:5px;border-top:0;border-right:0;border-bottom:0;border-left:0" width="100%">
                                                    <div class="spacer_block" style="height:40px;line-height:40px;font-size:1px"> </div>
                                                </td>
                                            </tr>
                                        </tbody>
                                    </table>
                                </td>
                            </tr>
                        </tbody>
                    </table>
                    <table class="row row-10" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0" width="100%" cellspacing="0" cellpadding="0" border="0" align="center">
                        <tbody>
                            <tr>
                                <td>
                                    <table class="row-content stack" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0;color:#000;width:680px" width="680" cellspacing="0" cellpadding="0" border="0" align="center">
                                        <tbody>
                                            <tr>
                                                <td class="column" style="mso-table-lspace:0;mso-table-rspace:0;font-weight:400;text-align:left;vertical-align:top;padding-top:5px;padding-bottom:5px;border-top:0;border-right:0;border-bottom:0;border-left:0" width="100%">
                                                    <table class="icons_block" role="presentation" style="mso-table-lspace:0;mso-table-rspace:0" width="100%" cellspacing="0" cellpadding="0" border="0">
                                                        <tbody><tr>
                                                            <td style="color:#9d9d9d;font-family:inherit;font-size:15px;padding-bottom:5px;padding-top:5px;text-align:center">
                                                                <table role="presentation" style="mso-table-lspace:0;mso-table-rspace:0" width="100%" cellspacing="0" cellpadding="0">
                                                                    <tbody><tr>
                                                                        <td style="text-align:center">
                                                                            <!--[if vml]><table align="left" cellpadding="0" cellspacing="0" role="presentation" style="display:inline-block;padding-left:0px;padding-right:0px;mso-table-lspace: 0pt;mso-table-rspace: 0pt;"><![endif]-->
                                                                            <!--[if !vml]><!-->
                                                                            <table class="icons-inner" style="mso-table-lspace:0;mso-table-rspace:0;display:inline-block;margin-right:-4px;padding-left:0;padding-right:0" role="presentation" cellspacing="0" cellpadding="0">
                                                                                <!--<![endif]-->
                                                                                <tbody><tr>
                                                                                    <td style="text-align:center;padding-top:5px;padding-bottom:5px;padding-left:5px;padding-right:6px">
                                                                                        <a href="https://www.designedwithbee.com/"><img class="icon" alt="Designed with BEE" src="https://d15k2d11r6t6rl.cloudfront.net/public/users/Integrators/BeeProAgency/53601_510656/Signature/bee.png" style="display:block;height:auto;border:0" width="34" height="32" align="middle"></a>
                                                                                    </td>
                                                                                    <td style="font-family:Arial,Helvetica Neue,Helvetica,sans-serif;font-size:15px;color:#9d9d9d;vertical-align:middle;letter-spacing:undefined;text-align:center">
                                                                                        <a href="https://www.designedwithbee.com/" style="color:#9d9d9d;text-decoration:none">Designed
                                                                                            with BEE</a></td>
                                                                                </tr>
                                                                            </tbody></table>
                                                                        </td>
                                                                    </tr>
                                                                </tbody></table>
                                                            </td>
                                                        </tr>
                                                    </tbody></table>
                                                </td>
                                            </tr>
                                        </tbody>
                                    </table>
                                </td>
                            </tr>
                        </tbody>
                    </table>
                </td>
            </tr>
        </tbody>
    </table><!-- End -->


</body></html>"""
    
    
    kirimEmail(email,subject,bodyFinish)
   
    return jsonify({})

@views.route('/approveworkingadmin', methods=['POST'])
def approveWorkingAdmin():
    data = json.loads(request.data)
    print(data)
    permitId = data['id']
    permit = Permit.query.filter_by(id=permitId).first()
    permit.status = 'waitinghost'
    db.session.commit()
    return jsonify({})

@views.route('/approveovertimeadmin', methods=['POST'])
def approveOvertimeAdmin():
    data = json.loads(request.data)
    print(data)
    permitId = data['id']
    permit = Permit.query.filter_by(id=permitId).first()
    permit.status = 'approved'
    db.session.commit()
    return jsonify({})

@views.route('/getcheckindata', methods=['POST'])
def getcheckindata():
    data = json.loads(request.data)
    print(':' in data)
    dataQ = {} 
    if ':' in data['qr']:

        qr = data['qr'].split(':')
        print(qr)
        nik = qr[1]
        permitId = qr[0]
        
        transaksi = Transaksi.query.filter_by(nik=nik).order_by(text('id desc')).first()

        permit = Permit.query.filter_by(id=permitId).first()

        
        if (nik in str(permit.anggota)) and (permit.status =='approved') :
            
            visitor = Visitor.query.filter_by(nik=nik).first()
            Host = permit.host.split(':')
            empID = Host[1]
            user = User.query.filter_by(empID=empID).first()
            print(visitor)
            if transaksi:
                if transaksi.status == 'checkout':
                    badge = Badge.query.filter_by(status='free').order_by(text('id asc')).first()

                    print(badge)
                    
                    dataQ['nama'] = visitor.nama
                    dataQ['nik'] = visitor.nik
                    dataQ['purpose'] = permit.purpose
                    dataQ['company'] = permit.namaVendor
                    dataQ['host'] = user.empID
                    dataQ['statusPermit'] = permit.status 
                
                    try:
                        dataQ['badge'] = badge.no
                    except:
                        dataQ['badge'] = '0'
                    dataQ['photo'] = visitor.photo
                    dataQ['qr'] = data['qr']
                    print(data['qr'])
            
                    return jsonify(dataQ)
                else:

                    dataT ={}
                    dataT['status'] = transaksi.status
                    dataT['nik'] = nik
                    
                    return jsonify(dataT)
            badge = Badge.query.filter_by(status='free').order_by(text('id asc')).first()

            print(badge)
            
            dataQ['nama'] = visitor.nama
            dataQ['nik'] = visitor.nik
            dataQ['purpose'] = permit.purpose
            dataQ['company'] = permit.namaVendor
            dataQ['host'] = user.empID
            dataQ['statusPermit'] = permit.status 
           
            try:
                dataQ['badge'] = badge.no
            except:
                dataQ['badge'] = '0'
            dataQ['photo'] = visitor.photo
            dataQ['qr'] = data['qr']
            print(data['qr'])
    
    return jsonify(dataQ)

@views.route('/savetotransaksicheckin', methods=['POST'])
def savetotransaksicheckin():
    data = json.loads(request.data)
       
    print(data)

    qr = data['qr'].split(':')
    print(qr)
    permitId = qr[0]
    nik = qr[1]
    no = data['badge']
    visitor = Visitor.query.filter_by(nik=nik).first()
    badge = Badge.query.filter_by(no=no).first()
    badge.status = 'used'
    permit = Permit.query.filter_by(id=permitId).first()
    dataTransaksi = {}
    dataTransaksi['namaVisitor'] = visitor.nama
    dataTransaksi['nik'] = nik
    dataTransaksi['purpose'] = permit.purpose
    dataTransaksi['vendor'] = permit.namaVendor
    dataTransaksi['host'] = permit.host
    dataTransaksi['timeCheckin'] = data['time']
    dataTransaksi['statusPermit'] = permit.status
    dataTransaksi['badge'] = data['badge']
    dataTransaksi['status'] = 'waiting'
    dataTransaksi['location'] = permit.location
    dataTransaksi['photo'] = visitor.photo
    # dataTransaksi['idPermit'] = permitId
    transaksi = Transaksi(**dataTransaksi)
    db.session.add(transaksi)
    
    db.session.commit()
    return jsonify({})

@views.route('/getcheckoutdatatransaksi', methods=['POST'])
def getcheckoutdatatransaksi():
    data = json.loads(request.data)
    print(data)
    badge = Badge.query.filter_by(rfid=data['rfid']).first()
    
    if badge:
        transaksi = Transaksi.query.filter_by(status='checkin',badge=badge.no).first()
        if transaksi:
                dataQ = {}
                dataQ['id'] = transaksi.id
                dataQ['namaVisitor'] = transaksi.namaVisitor
                dataQ['nik'] = transaksi.nik
                dataQ['purpose'] = transaksi.purpose
                dataQ['vendor'] = transaksi.vendor
                dataQ['host'] = transaksi.host
                dataQ['statusPermit'] = transaksi.statusPermit
                dataQ['badge'] = transaksi.badge
                dataQ['status'] = transaksi.status
                dataQ['location'] = transaksi.location
                dataQ['photo'] = transaksi.photo
                return jsonify(dataQ)
        else:
            return jsonify({})
 

    else:
        return jsonify({})


@views.route('/updatedatacheckout', methods=['POST'])
def updatedatacheckout():
    data = json.loads(request.data)
    nik = data['Nik']
    transaksi = Transaksi.query.filter_by(nik=nik).first()
    tco = data['TimeCO']
    transaksi.timeCheckout = tco
    transaksi.status = 'checkout'
    badgeVis = data['Badge']
    badge = Badge.query.filter_by(status='used',no=badgeVis).first()
    badge.status = 'free'

    db.session.commit()
    print("data checkout updated")
    return jsonify({})

@views.route('/export', methods=['POST'])
def export():
    data = json.loads(request.data)
    by = data['by']
    transaksi = Transaksi.query.order_by(by+" desc").all()
    data = {} 
    data['id'] = transaksi.id
    data['namaVisitor'] = transaksi.namaVisitor
    data['nik'] = transaksi.nik
    data['purpose'] = transaksi.purpose
    data['vendor'] = transaksi.vendor
    data['host'] = transaksi.host
    data['timeChecout'] = datetime.datetime.now()
    data['statusPermit'] = transaksi.statusPermit
    data['badge'] = transaksi.badge
    data['status'] = transaksi.status
    data['location'] = transaksi.location
    data['photo'] = transaksi.photo


    return jsonify(data)

@views.route('/exportxls', methods=['POST'])
def exportxls():
    data = json.loads(request.data)
    By = data['by']
    order = By + ' desc'
    transaksi = Transaksi.query.order_by(order).all()
    dataQ = {} 
    dataQ['id'] = transaksi.id
    return jsonify(dataQ)
    # dataQ['namaVisitor'] = transaksi.namaVisitor
    # dataQ['nik'] = transaksi.nik
    # dataQ['purpose'] = transaksi.purpose
    # dataQ['vendor'] = transaksi.vendor
    # dataQ['host'] = transaksi.host
    # dataQ['timeChecout'] = datetime.datetime.now()
    # dataQ['statusPermit'] = transaksi.statusPermit
    # dataQ['badge'] = transaksi.badge
    # dataQ['status'] = transaksi.status
    # dataQ['location'] = transaksi.location
    # dataQ['photo'] = transaksi.photo
    # # for i in transaksi:
    #     print(len(i))    

    # print(dataQ)


    