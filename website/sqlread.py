import sqlalchemy as db
from sqlalchemy import create_engine, select, table, column,text,insert
import json
import openpyxl
from openpyxl import load_workbook


# dbpana = "mssql+pymssql://sa:Batam2021@103.142.240.134:1433/VMS"
dbpana = "mssql+pymssql://sa:Batam2021@10.89.1.50:1433/VMS"

def readsql():
    engine = db.create_engine(dbpana)
    connection = engine.connect()
    metadata = db.MetaData()
    census = db.Table('permit', metadata, autoload=True, autoload_with=engine)

    #Equivalent to 'SELECT * FROM census'
    query = db.select([census])

    ResultProxy = connection.execute(query)

    ResultSet = ResultProxy.fetchall()


    dictDb = {}
    listDb =[]
    for i in ResultSet:
        dictDb['no'] = list(i)[0]


        dictDb['subDate'] = list(i)[1]

        dictDb['namaVendor'] = list(i)[2]
        dictDb['startDate'] = list(i)[3]
        dictDb['endDate'] = list(i)[4]
        dictDb['purpose'] = list(i)[5]
        dictDb['location'] = list(i)[6]
        dictDb['supplyBarang'] = list(i)[7]
        dictDb['permitNo'] = list(i)[8]
        dictDb['desk'] = list(i)[9]
        dictDb['anggota'] = list(i)[10]
        dictDb['email'] = list(i)[11]
        dictDb['barangBawaan'] = list(i)[12]
        dictDb['sign'] = list(i)[13]




        listDb.append(dictDb.copy())

    print(listDb)

# readsql()

def findDatapermit(cari):
    engine = db.create_engine(dbpana)
    connection = engine.connect()
    metadata = db.MetaData()
    census = db.Table('permit', metadata, autoload=True, autoload_with=engine)

    #Equivalent to 'SELECT * FROM census'
    query = db.select([census])

    ResultProxy = connection.execute(query)

    ResultSet = ResultProxy.fetchall()
    permitNo = []
    for i in ResultSet:
        permitNo.append(i[8])
    noData = permitNo.index(cari)

    print(list(ResultSet[noData]))

#readsql()

#findDatapermit('545466791')


def getNikVisitor(nik):
    nikVisitor = {}
    engine = create_engine("mssql+pymssql://sa:Batam2021@103.142.240.134:1433/VMS",future=True)
    with engine.connect() as conn:
        
        hasil = conn.execute(text(f"SELECT * FROM visitor WHERE nik='{nik}'"))

        for h in hasil:
            nikVisitor = dict(h)

    print(nikVisitor)

# getNikVisitor('834792873489')

def genPermitXLS(id):
    nikVisitor = {}
    engine = create_engine("mssql+pymssql://sa:Batam2021@103.142.240.134:1433/VMS",future=True)
    with engine.connect() as conn:

    # nikVisitor;
        
        hasil = conn.execute(text(f"SELECT * FROM permit WHERE id='{id}'"))

        for h in hasil:
            data = dict(h)

    # print(data)

    sourcefile = r'./static/VisitorApprovalBT.xlsx';

    wb = load_workbook(sourcefile);

    # sheet = wb['Visitor Approval'];
    # sheet['C5'] = data['namaVendor']
    # # sheet['C6'] = design
    # sheet['C7'] = data['namaVendor']
    # sheet['C9'] = data['subDate']
    # sheet['C10'] = data['startDate']
    # sheet['E10'] = data['endDate']
    # sheet['A11'] = purpose['pupose']

    # barang = json.loads(data['barangBawaan'])
    

    ws = wb['Visitor Approval']
    locations = []
    for row in ws.iter_rows(min_row=36, min_col=3, max_col=3, max_row=65, values_only=True):
        x = locations.append(row[0])

    detaillocations = {}
    for i, v in enumerate(locations, start=36):
        detaillocations[i] = v

    print(detaillocations)
        
    # wb.save(sourcefile)

    # return sourcefile
# genPermitXLS(15)




datapermit = [{'no': 1, 'subDate': datetime.datetime(2021, 11, 12, 14, 50, 55), 'namaVendor': 'CV. ANINDYATRANS ', 'startDate': datetime.datetime(2021, 11, 23, 8, 0), 'endDate': datetime.datetime(2021, 11, 23, 13, 59), 'purpose': 'SUPPLY', 'location': 'Workshop & Machine shop', 'supplyBarang': '[{"Item Description":"Buku","Quantity":"3"},{"Item Description":"pulpen","Quantity":"100"}]', 'permitNo': '', 'desk': '', 'anggota': '[{"Nama": "Faris", "Jabatan": "Manager", "NIK": "23784205723407", "Register": "tidak", "Covid": "tidak"}]', 'email': 'blgr182@gmail.com', 'barangBawaan': 'UMI', 'sign': 'TIDAK'}, {'no': 2, 'subDate': datetime.datetime(2021, 11, 11, 11, 50, 52), 'namaVendor': 'PT. AIK MOH CHEMICALS INDONESIA', 'startDate': datetime.datetime(2021, 11, 13, 8, 0), 'endDate': datetime.datetime(2021, 11, 18, 13, 59), 'purpose': 'WORKING', 'location': 'Resistor production', 'supplyBarang': '', 'permitNo': '5647754647', 'desk': 'fghdfhgdfhd', 'anggota': '[{"Nama": "Novel", "Jabatan": "Komisaris", "NIK": "83423849028", "Register": "tidak", "Covid": "tidak"}, {"Nama": "Faris", "Jabatan": "Manager", "NIK": "982982897873", "Register": "tidak", "Covid": "tidak"}]', 'email': 'novel@martin.harianto', 'barangBawaan': 'FITRIANSYAH', 'sign': 'YA'}, {'no': 3, 'subDate': datetime.datetime(2021, 11, 12, 14, 50, 55), 'namaVendor': 'CV. ANINDYATRANS ', 'startDate': datetime.datetime(2021, 11, 23, 8, 0), 'endDate': datetime.datetime(2021, 11, 23, 13, 59), 'purpose': 'SUPPLY', 'location': 'LOGISTIC', 'supplyBarang': '[{"Item Description":"Buku","Quantity":"3"},{"Item Description":"pulpen","Quantity":"100"}]', 'permitNo': '', 'desk': '', 'anggota': '[{"Nama": "Faris", "Jabatan": "Manager", "NIK": "23784205723407", "Register": "tidak", "Covid": "tidak"}]', 'email': 'blgr182@gmail.com', 'barangBawaan': 'UMI', 'sign': 'TIDAK'}, {'no': 4, 'subDate': datetime.datetime(2021, 11, 11, 11, 50, 52), 'namaVendor': 'PT. AIK MOH CHEMICALS INDONESIA', 'startDate': datetime.datetime(2021, 11, 13, 8, 0), 'endDate': datetime.datetime(2021, 11, 18, 13, 59), 'purpose': 'WORKING', 'location': 'Workshop & Machine shop', 'supplyBarang': '', 'permitNo': '5647754647', 'desk': 'fghdfhgdfhd', 'anggota': '[{"Nama": "Novel", "Jabatan": "Komisaris", "NIK": "83423849028", "Register": "tidak", "Covid": "tidak"}, {"Nama": "Faris", "Jabatan": "Manager", "NIK": "982982897873", "Register": "tidak", "Covid": "tidak"}]', 'email': 'novel@martin.harianto', 'barangBawaan': 'FITRIANSYAH', 'sign': 'YA'}, {'no': 5, 'subDate': datetime.datetime(2021, 11, 12, 14, 50, 55), 'namaVendor': 'CV. ANINDYATRANS ', 'startDate': datetime.datetime(2021, 11, 23, 8, 0), 'endDate': datetime.datetime(2021, 11, 23, 13, 59), 'purpose': 'SUPPLY', 'location': 'IT Server Room', 'supplyBarang': '[{"Item Description":"Buku","Quantity":"3"},{"Item Description":"pulpen","Quantity":"100"}]', 'permitNo': '', 'desk': '', 'anggota': '[{"Nama": "Faris", "Jabatan": "Manager", "NIK": "23784205723407", "Register": "tidak", "Covid": "tidak"}]', 'email': 'blgr182@gmail.com', 'barangBawaan': 'UMI', 'sign': 'TIDAK'}, {'no': 6, 'subDate': datetime.datetime(2021, 11, 11, 11, 50, 52), 'namaVendor': 'PT. AIK MOH CHEMICALS INDONESIA', 'startDate': datetime.datetime(2021, 11, 13, 8, 0), 'endDate': datetime.datetime(2021, 11, 18, 13, 59), 'purpose': 'WORKING', 'location': 'Resistor Office', 'supplyBarang': '', 'permitNo': '5647754647', 'desk': 'fghdfhgdfhd', 'anggota': '[{"Nama": "Novel", "Jabatan": "Komisaris", "NIK": "83423849028", "Register": "tidak", "Covid": "tidak"}, {"Nama": "Faris", "Jabatan": "Manager", "NIK": "982982897873", "Register": "tidak", "Covid": "tidak"}]', 'email': 'novel@martin.harianto', 'barangBawaan': 'FITRIANSYAH', 'sign': 'YA'}]